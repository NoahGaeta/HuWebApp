from GoogleDriveHandler import GoogleDriveHandler
from copy import copy
import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook


class LocalExcelSheetIntegrator:

    VALID_COLUMNS = [
        'Litter No/Pup No', 'Animal ID', 'Sex', 'Genotype', 'Strain', 'Set Up date',
        'Cage #', 'Destination', 'Generation', 'Line', 'DOB', 'Weaning Date',
        'Genotyping Date', 'Set Up date', 'Comments', 'Source', 'Father', 'Mother', 'Mother ', 'Genotyping date'
    ]

    INVALID_COLUMNS = [
        'Parents'
    ]

    def __init__(self, user_file):
        self.user_file = user_file
        self.old_work_book = load_workbook(self.user_file)
        self.work_book = Workbook()
        del self.work_book['Sheet']
        self.max_row = 1
        self.animal_id = []
        self.headers_inserted = False

    def create_new_excel_sheet(self, master_sheet_name):
        master_sheet = self.work_book.create_sheet(master_sheet_name)
        for sheet_name in self.old_work_book.sheetnames:
            o = self.work_book.create_sheet(sheet_name)
            safe_title = o.title
            if master_sheet != sheet_name:
                self.__copy_sheet(self.old_work_book[sheet_name], self.work_book[safe_title])
                self.__get_all_valid_columns(self.old_work_book[sheet_name], master_sheet)

        self.work_book.save(self.user_file)

    def __get_all_valid_columns(self, source_sheet, master_sheet):
        columns = source_sheet.columns

        for items in sorted(source_sheet.merged_cell_ranges):
            source_sheet.unmerge_cells(str(items))

        found_header_row = False

        row = self.max_row
        temp_max_row = self.max_row
        ignored_row = []
        ignored_column = []
        header_map = {}
        header_row_num = None

        for column in columns:
            for cell in column:
                if type(cell.value) == datetime.datetime:
                    cell.value = str(cell.value)

                if cell.value in self.animal_id:
                    ignored_row.append(row)
                    print(cell.value)

                if found_header_row and row not in ignored_row and cell.value not in self.VALID_COLUMNS and cell.value not in self.animal_id and cell.col_idx not in ignored_column:
                    master_sheet.cell(row=row + 1, column=cell.col_idx, value=cell.value)
                    row += 1

                if header_map.get(cell.col_idx) == 'Animal ID' and cell.value not in self.VALID_COLUMNS and cell.value and cell.value not in self.animal_id:
                    self.animal_id.append(cell.value)

                if cell.value in self.VALID_COLUMNS or cell.value in self.INVALID_COLUMNS:
                    found_header_row = True
                    header_row_num = row

                if found_header_row and cell.value in self.VALID_COLUMNS and header_row_num == row:
                    header_map.update({header_row_num: cell.value})

                if cell.value in self.INVALID_COLUMNS:
                    ignored_column.append(cell.col_idx)

            self.max_row = row
            row = temp_max_row

        self.headers_inserted = True

    def __validate_cell_value(self):
        pass

    def __copy_sheet(self, source_sheet, new_sheet):
        columns = source_sheet.columns
        for items in sorted(source_sheet.merged_cell_ranges):
            source_sheet.unmerge_cells(str(items))
        for column in columns:
            for cell in column:
                if type(cell.value) == datetime.datetime:
                    cell.value = str(cell.value)
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                self.__format_cell(cell, new_cell)

    def __format_cell(self, cell, new_cell):
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
        return new_cell


class GoogleExcelSheetIntegrator:

    def __init__(self):
        self.google_drive = GoogleDriveHandler()

    def handle_file_input_google_drive(self, user_file_list, master_file):
        file_obj = self.google_drive.find_file(master_file)
        temporary_file = self.google_drive.download_google_document(file_obj, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '.xls')
        file_path_list = [temporary_file]
        for user_file in user_file_list:
            file_path_list.append(user_file)
        self.__combine_sheets_seperate_sheets(file_path_list, temporary_file)
        return self.google_drive.update_file(file_obj, temporary_file)

    def __combine_sheets_seperate_sheets(self, file_path_list, export_path):
        print("Merging Excel Sheets...")
        excel_sheet_list = []

        for file_path in file_path_list:
            excel_sheet_list.append(pd.read_excel(file_path))

        final_dataframe = pd.concat(excel_sheet_list)

        final_dataframe.to_excel(export_path, index=False)
